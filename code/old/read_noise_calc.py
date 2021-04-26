#!/usr/bin/env python
# coding: utf-8
# Denis Varise Bernardes.
# 08/10/2019.

"""
This software has the ReadNoiseCalc class. This class calculates the read noise of the SPARC4 EMCCDs as a function of their
operation mode. The calculations are done based on a series of characterization of the SPARC4 cameras. For the
conventional mode, it is read the respective value of the read noise in the Tabelas_Valores_Ruido_Leitura spreadsheet.
For the EM mode, it is done an interpolation of the data presented by the respective spreadshhet, as a function
of the EM gain.
"""

import openpyxl
import numpy as np
from scipy.interpolate import interp1d
from sys import exit


class ReadNoiseCalc:

    def __init__(self):
        self.noise = 0

    def write_operation_mode(self, em_mode, em_gain, hss, preamp, binn):
        # This function writes the CCD operation mode to the class
        self.em_mode = em_mode
        self.em_gain = em_gain
        self.hss = hss
        self.preamp = preamp
        self.binn = binn
        self.read_noise = 0

    def get_operation_mode(self):
        print('em_mode = ', self.em_mode)
        print('em_gain = ', self.em_gain)
        print('hss = ', self.hss)
        print('preamp = ', self.preamp)
        print('binn = ', self.binn)

    def calc_read_noise(self):
        # This function calculates the read noise of the CCD as a function of its operation mode
        read_noise = 0
        if self.em_mode == 0:
            # For the conventional mode, it is used the read noise values of the Tabelas_Valores_Ruido_Leitura spreadsheet
            indice_tab = 0
            if self.hss == 1:
                if self.preamp == 1:
                    if self.binn == 1:
                        indice_tab = 19
                    if self.binn == 2:
                        indice_tab = 20
                if self.preamp == 2:
                    if self.binn == 1:
                        indice_tab = 21
                    if self.binn == 2:
                        indice_tab = 22
            if self.hss == 0.1:
                if self.preamp == 1:
                    if self.binn == 1:
                        indice_tab = 23
                    if self.binn == 2:
                        indice_tab = 24
                if self.preamp == 2:
                    if self.binn == 1:
                        indice_tab = 25
                    if self.binn == 2:
                        indice_tab = 26
            spreadsheet = openpyxl.load_workbook(
                r'spreadsheet\Read_noise_and_gain_values.xlsx').active
            read_noise = spreadsheet.cell(indice_tab, 6).value

        if self.em_mode == 1:
            # For the EM mode, the read noise is obtained through an interpolation of
            # the values presente by the respective spreadsheet, as a function of the CCD EM gain.
            tab_name = 'spreadsheet/RN_PA' + \
                str(int(self.preamp)) + 'B' + str(int(self.binn)) + \
                'HSS' + str(int(self.hss)) + '.xlsx'
            spreadsheet = list(openpyxl.load_workbook(tab_name).active.values)
            column_em_gain = [value[0] for value in spreadsheet[1:12]]
            column_noise = [value[1] for value in spreadsheet[1:12]]
            f = interp1d(column_em_gain, column_noise)
            read_noise = f(self.em_gain)

        self.read_noise = float(read_noise)

        return self.read_noise
