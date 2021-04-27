# -*- coding: utf-8 -*-
"""
Read Noise Calculation Class
============================

This software has the ReadNoiseCalc class. This class calculates the read noise
of the SPARC4 EMCCDs as a function of their operation mode. The calculations
are done based on a series of characterization of the SPARC4 cameras. For the
conventional mode, it is read the respective value of the read noise in the
Tabelas_Valores_Ruido_Leitura spreadsheet. For the EM mode, it is done an
interpolation of the data presented by the respective spreadshhet, as a
function of the EM gain.
"""

# Denis Varise Bernardes.
# 08/10/2019.

import openpyxl
from scipy.interpolate import interp1d


class Read_Noise_Calculation:
    """Read Noise Calculation Class.

    Parameters
    ----------
    ccd_operation_mode: dictionary

        A dictionary with the parameter of the CCD operation mode.
        em_mode : [0, 1]
            CCD Electron Multiplying Mode
        em_gain : float
            CCD Electron Multiplying gain
        hss : [0.1, 1, 10, 20, 30]
            Horizontal Shift Spedd of the pixels
        preamp : [1, 2]
            Pre-amplifer gain
        binn : [1, 2]
            Binning of the pixels

    directory : string
        Directory of the spreadsheet with the read noise of the CCD
        """

    def __init__(self, ccd_operation_mode, directory):
        """Initialize the class."""
        self.em_mode = ccd_operation_mode['em_mode']
        self.em_gain = ccd_operation_mode['em_gain']
        self.hss = ccd_operation_mode['hss']
        self.preamp = ccd_operation_mode['preamp']
        self.binn = ccd_operation_mode['binn']
        self.directory = directory

    def get_operation_mode(self):
        """Print the operation mode on the screen."""
        print('em_mode = ', self.em_mode)
        print('em_gain = ', self.em_gain)
        print('hss = ', self.hss)
        print('preamp = ', self.preamp)
        print('binn = ', self.binn)

    def calculate_read_noise(self):
        """Calculate the read noise of the CCD.

        For the conventional mode, it is used the read noise values of the
        Read_noise_and_gain_values spreadsheet

        For the EM mode, the read noise is obtained through an interpolation of
        the values presente by the respective spreadsheet, as a function of the
        CCD EM gain.
        """
        if self.em_mode == 0:
            self._calculate_read_noise_conventional_mode()
        if self.em_mode == 1:
            self._calculate_read_noise_em_mode()
        return self.read_noise

    def _calculate_read_noise_conventional_mode(self):
        """Calculate the read noise for the conventional mode."""
        indice_tab = 0
        if self.hss == 1:
            if self.preamp == 1:
                indice_tab = 19
            if self.preamp == 2:
                indice_tab = 21
        if self.hss == 0.1:
            if self.preamp == 1:
                indice_tab = 23
            if self.preamp == 2:
                indice_tab = 25
        if self.binn == 2:
            indice_tab += 1
        path = r'RNC/spreadsheet' \
            + '/' + self.directory + '/' + 'Read_noise_and_gain_values.xlsx'

        spreadsheet = openpyxl.load_workbook(path).active
        self.read_noise = spreadsheet.cell(indice_tab, 6).value

    def _calculate_read_noise_em_mode(self):
        """Calculate the read noise for the EM mode."""
        tab_name = 'RNC/spreadsheet' + '/' + self.directory + '/' + 'RN_PA'\
            + str(int(self.preamp)) + 'B' + str(int(self.binn))\
            + 'HSS' + str(int(self.hss)) + '.xlsx'
        spreadsheet = list(openpyxl.load_workbook(tab_name).active.values)
        column_em_gain = [value[0] for value in spreadsheet[1:12]]
        column_noise = [value[1] for value in spreadsheet[1:12]]
        f = interp1d(column_em_gain, column_noise)
        read_noise = f(self.em_gain)

        self.read_noise = float(read_noise)
